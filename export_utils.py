"""
Data Export Utilities
Export predictions as CSV, PDF, or Excel formats.
"""

import pandas as pd
from io import BytesIO, StringIO
from datetime import datetime
from typing import List, Dict
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_to_csv(predictions: List[Dict], filename: str = None) -> bytes:
    """
    Export predictions to CSV.
    
    Args:
        predictions: List of prediction dictionaries
        filename: Optional filename
    
    Returns:
        CSV bytes
    """
    df = pd.DataFrame(predictions)
    
    # Format columns
    if "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"]).dt.strftime("%Y-%m-%d %H:%M:%S")
    
    csv_buffer = StringIO()
    df.to_csv(csv_buffer, index=False)
    
    return csv_buffer.getvalue().encode()


def export_to_excel(predictions: List[Dict], user_stats: Dict = None, filename: str = None) -> bytes:
    """
    Export predictions to Excel with formatting.
    
    Args:
        predictions: List of prediction dictionaries
        user_stats: Optional user statistics
        filename: Optional filename
    
    Returns:
        Excel bytes
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws_data = wb.active
    ws_data.title = "Predictions"
    
    # Create stats sheet if provided
    if user_stats:
        ws_stats = wb.create_sheet("Statistics")
        _create_stats_sheet(ws_stats, user_stats)
    
    # Create data sheet
    _create_data_sheet(ws_data, predictions)
    
    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()


def _create_stats_sheet(ws, stats: Dict):
    """Create statistics sheet."""
    # Header
    ws["A1"] = "User Statistics"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    
    # Stats
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    ws["A3"].font = Font(bold=True)
    ws["B3"].font = Font(bold=True)
    
    row = 4
    for key, value in stats.items():
        ws[f"A{row}"] = key.replace("_", " ").title()
        ws[f"B{row}"] = value if not isinstance(value, float) else round(value, 2)
        row += 1
    
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20


def _create_data_sheet(ws, predictions: List[Dict]):
    """Create data sheet with formatting."""
    if not predictions:
        ws["A1"] = "No predictions found"
        return
    
    df = pd.DataFrame(predictions)
    
    # Headers
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column.replace("_", " ").title()
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # Format numeric columns
            if isinstance(value, float):
                cell.value = round(value, 2)
                cell.number_format = "0.00"
            
            cell.alignment = Alignment(horizontal="right")
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    # Auto-adjust column widths
    for col_idx, column in enumerate(df.columns, 1):
        max_length = min(25, max(
            len(str(df.iloc[:, col_idx - 1].max())),
            len(column)
        ) + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length


def export_to_pdf(
    predictions: List[Dict],
    user_stats: Dict = None,
    username: str = "User",
    filename: str = None
) -> bytes:
    """
    Export predictions to PDF with formatting.
    
    Args:
        predictions: List of prediction dictionaries
        user_stats: Optional user statistics
        username: Username for header
        filename: Optional filename
    
    Returns:
        PDF bytes
    """
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor("#2E7D32"),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor("#2E7D32"),
        spaceAfter=10,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    # Title
    elements.append(Paragraph("🌾 Crop Yield Prediction Report", title_style))
    elements.append(Paragraph(f"User: <b>{username}</b> | Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}", 
                             styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # User Statistics
    if user_stats:
        elements.append(Paragraph("User Statistics", section_style))
        stats_data = [["Metric", "Value"]]
        for key, value in user_stats.items():
            metric_name = key.replace("_", " ").title()
            value_str = f"{value:.2f}" if isinstance(value, float) else str(value)
            stats_data.append([metric_name, value_str])
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f0f0")]),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Predictions
    if predictions:
        elements.append(Paragraph("Prediction History", section_style))
        
        # Create table data
        if predictions:
            df = pd.DataFrame(predictions)
            # Select key columns only
            key_columns = ['timestamp', 'crop', 'state', 'predicted_yield', 'total_profit']
            df_subset = df[[col for col in key_columns if col in df.columns]]
            
            table_data = [[col.replace("_", " ").title() for col in df_subset.columns]]
            for _, row in df_subset.iterrows():
                table_data.append([str(v)[:25] for v in row.values])
            
            pred_table = Table(table_data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1.2*inch, 1.2*inch])
            pred_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f0f0")]),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(pred_table)
    
    # Build PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    return pdf_buffer.getvalue()


def get_export_filename(export_type: str, username: str = "export") -> str:
    """Generate filename for export."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    extensions = {"csv": "csv", "excel": "xlsx", "pdf": "pdf"}
    return f"{username}_predictions_{timestamp}.{extensions.get(export_type, 'csv')}"
